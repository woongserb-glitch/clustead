import csv
import re
from collections import Counter
from pathlib import Path
from xml.etree import ElementTree
from zipfile import ZipFile

from transaction_layer_utils import (
    MOLIT_TRANSACTION_RAW_DIR,
    TRANSACTION_MASTER_PATH,
    clean_text,
    compose_lot_number,
    ensure_transaction_dirs,
    first_value,
    normalize_address,
    normalize_lot_part,
    normalize_money,
    normalize_name,
    parse_float,
    parse_int,
    safe_read_text_bytes,
    write_csv,
    year_month_from_date,
)


MASTER_FIELDS = [
    "source",
    "source_file",
    "transaction_type",
    "deal_type",
    "sido",
    "gu",
    "dong",
    "legal_dong",
    "jibun",
    "bonbun",
    "bubun",
    "road_name",
    "road_address",
    "normalized_road_address",
    "apartment_name",
    "normalized_apartment_name",
    "contract_date",
    "contract_year",
    "contract_month",
    "area_m2",
    "floor",
    "built_year",
    "trade_price_manwon",
    "deposit_manwon",
    "monthly_rent_manwon",
    # Compatibility fields consumed by existing mapping/summary/runtime code.
    "deal_date",
    "deal_year",
    "deal_month",
    "apt_name_raw",
    "apt_name_normalized",
    "build_year",
    "deal_amount",
    "deposit_amount",
    "monthly_rent",
    "lat",
    "lng",
    "source_year",
    "building_use",
]

HEADER_REQUIRED_TOKENS = ("시군구", "단지명", "계약년월")


def detect_transaction_type(path, headers):
    name = path.name.lower()
    header_text = " ".join(headers)
    if any(token in name for token in ("rent", "jeonse", "월세", "전세", "전월세")):
        return "rent"
    if any(token in name for token in ("trade", "sale", "매매", "실거래")):
        return "trade"
    if "보증금" in header_text or "월세" in header_text:
        return "rent"
    return "trade"


def detect_year(path, row):
    match = re.search(r"(20\d{2})", path.name)
    if match:
        return match.group(1)
    ym = clean_text(row.get("계약년월"))
    return ym[:4]


def normalize_header(value):
    return re.sub(r"\s+", "", clean_text(value))


def trim_empty_tail(values):
    items = list(values)
    while items and not clean_text(items[-1]):
        items.pop()
    return items


def find_header_index(rows):
    for index, row in enumerate(rows):
        normalized = [normalize_header(value) for value in row]
        if all(token in normalized for token in HEADER_REQUIRED_TOKENS):
            return index
    if len(rows) >= 13:
        return 12
    return -1


def rows_from_xlsx(path):
    try:
        import openpyxl
    except ImportError:
        return rows_from_xlsx_stdlib(path)

    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    worksheet = workbook[workbook.sheetnames[0]]
    if getattr(worksheet, "max_row", None) == 1 and getattr(worksheet, "max_column", None) == 1:
        # MOLIT exports sometimes declare only A1 in sheet dimensions.
        # Reset so openpyxl scans the actual rows before header detection.
        reset_dimensions = getattr(worksheet, "reset_dimensions", None)
        if reset_dimensions:
            reset_dimensions()
    raw_rows = [trim_empty_tail(row) for row in worksheet.iter_rows(values_only=True)]
    header_index = find_header_index(raw_rows)
    if header_index < 0:
        return rows_from_xlsx_stdlib(path)

    headers = [normalize_header(value) for value in raw_rows[header_index]]
    records = []
    for raw in raw_rows[header_index + 1:]:
        if not any(clean_text(value) for value in raw):
            continue
        row = {headers[i]: raw[i] if i < len(raw) else "" for i in range(len(headers))}
        if clean_text(row.get("NO")).upper() == "NO":
            continue
        records.append(row)
    return records, headers


def excel_column_to_index(cell_ref):
    letters = re.sub(r"[^A-Z]", "", clean_text(cell_ref).upper())
    index = 0
    for char in letters:
        index = index * 26 + (ord(char) - ord("A") + 1)
    return max(index - 1, 0)


def read_shared_strings(archive):
    try:
        root = ElementTree.fromstring(archive.read("xl/sharedStrings.xml"))
    except KeyError:
        return []
    namespace = {"x": "http://schemas.openxmlformats.org/spreadsheetml/2006/main"}
    values = []
    for item in root.findall("x:si", namespace):
        parts = [node.text or "" for node in item.findall(".//x:t", namespace)]
        values.append("".join(parts))
    return values


def rows_from_xlsx_stdlib(path):
    namespace = {"x": "http://schemas.openxmlformats.org/spreadsheetml/2006/main"}
    with ZipFile(path) as archive:
        shared_strings = read_shared_strings(archive)
        sheet_names = sorted(
            name for name in archive.namelist()
            if re.match(r"xl/worksheets/sheet\d+\.xml$", name)
        )
        if not sheet_names:
            raise RuntimeError(f"no worksheet found in MOLIT xlsx: {path}")
        root = ElementTree.fromstring(archive.read(sheet_names[0]))
        raw_rows = []
        for row_node in root.findall(".//x:sheetData/x:row", namespace):
            values = []
            for cell in row_node.findall("x:c", namespace):
                col_index = excel_column_to_index(cell.attrib.get("r", "A1"))
                while len(values) <= col_index:
                    values.append("")
                value_node = cell.find("x:v", namespace)
                inline_node = cell.find("x:is/x:t", namespace)
                raw_value = value_node.text if value_node is not None else (inline_node.text if inline_node is not None else "")
                if cell.attrib.get("t") == "s" and raw_value != "":
                    try:
                        raw_value = shared_strings[int(raw_value)]
                    except Exception:
                        pass
                values[col_index] = raw_value
            raw_rows.append(trim_empty_tail(values))

    header_index = find_header_index(raw_rows)
    if header_index < 0:
        raise RuntimeError(f"could not detect MOLIT header row: {path}")
    headers = [normalize_header(value) for value in raw_rows[header_index]]
    records = []
    for raw in raw_rows[header_index + 1:]:
        if not any(clean_text(value) for value in raw):
            continue
        row = {headers[i]: raw[i] if i < len(raw) else "" for i in range(len(headers))}
        records.append(row)
    return records, headers


def rows_from_csv(path):
    text, _ = safe_read_text_bytes(path.read_bytes())
    reader = csv.reader(text.splitlines())
    raw_rows = [trim_empty_tail(row) for row in reader]
    header_index = find_header_index(raw_rows)
    if header_index < 0:
        raise RuntimeError(f"could not detect MOLIT header row: {path}")
    headers = [normalize_header(value) for value in raw_rows[header_index]]
    records = []
    for raw in raw_rows[header_index + 1:]:
        if not any(clean_text(value) for value in raw):
            continue
        records.append({headers[i]: raw[i] if i < len(raw) else "" for i in range(len(headers))})
    return records, headers


def iter_molit_sources():
    ensure_transaction_dirs()
    for path in sorted(MOLIT_TRANSACTION_RAW_DIR.iterdir()):
        if path.suffix.lower() in (".xlsx", ".xlsm"):
            rows, headers = rows_from_xlsx(path)
            yield path, rows, headers
        elif path.suffix.lower() in (".csv", ".txt"):
            rows, headers = rows_from_csv(path)
            yield path, rows, headers


def split_sigungu(value):
    parts = clean_text(value).split()
    sido = parts[0] if parts else ""
    gu = parts[1] if len(parts) >= 2 else ""
    dong = " ".join(parts[2:]) if len(parts) >= 3 else ""
    return sido, gu, dong


def contract_date(row):
    ym = re.sub(r"\D", "", clean_text(first_value(row, ("계약년월", "년월"))))
    day = re.sub(r"\D", "", clean_text(first_value(row, ("계약일", "일"))))
    if len(ym) < 6:
        return ""
    if not day:
        day = "1"
    try:
        return f"{int(ym[:4]):04d}-{int(ym[4:6]):02d}-{int(day):02d}"
    except Exception:
        return ""


def road_address_for_row(sido, gu, road_name):
    road = clean_text(road_name)
    if not road or road == "-":
        return ""
    if " " in road and any(road.startswith(prefix) for prefix in ("서울특별시", "서울시")):
        return road
    return " ".join(part for part in (sido or "서울특별시", gu, road) if part)


def normalize_row(row, source_file, transaction_type):
    sido, gu, dong = split_sigungu(first_value(row, ("시군구", "법정동")))
    apartment_name = first_value(row, ("단지명", "아파트명"))
    road_name = first_value(row, ("도로명", "도로명주소"))
    bonbun = normalize_lot_part(first_value(row, ("본번",)))
    bubun = normalize_lot_part(first_value(row, ("부번",)))
    if bubun == "0":
        bubun = ""
    jibun = clean_text(first_value(row, ("번지", "지번"))) or compose_lot_number(bonbun, bubun)
    if not bonbun and jibun:
        parts = jibun.split("-", 1)
        bonbun = normalize_lot_part(parts[0])
        bubun = normalize_lot_part(parts[1] if len(parts) > 1 else "")
        if bubun == "0":
            bubun = ""

    date_value = contract_date(row)
    contract_year, contract_month = year_month_from_date(date_value)
    road_address = road_address_for_row(sido, gu, road_name)
    housing_type = clean_text(first_value(row, ("주택유형", "건물용도")))
    if housing_type and housing_type != "아파트":
        return None

    monthly = normalize_money(first_value(row, ("월세금(만원)", "월세금", "월세")))
    deposit = normalize_money(first_value(row, ("보증금(만원)", "보증금", "보증금액")))
    trade_price = normalize_money(first_value(row, ("거래금액(만원)", "거래금액", "매매금액")))
    if transaction_type == "trade":
        deal_type = "매매"
        if trade_price is None:
            return None
    else:
        monthly = monthly or 0
        deal_type = "월세" if monthly > 0 else "전세"
        if deposit is None:
            return None

    if not apartment_name or not date_value:
        return None

    area = parse_float(first_value(row, ("전용면적(㎡)", "전용면적", "전용면적(m2)")))
    floor = parse_int(first_value(row, ("층",)))
    built_year = parse_int(first_value(row, ("건축년도", "건축연도")))

    base = {
        "source": "molit_rt_xls",
        "source_file": source_file,
        "transaction_type": transaction_type,
        "deal_type": deal_type,
        "sido": sido,
        "gu": gu,
        "dong": dong,
        "legal_dong": dong,
        "jibun": jibun,
        "bonbun": bonbun,
        "bubun": bubun,
        "road_name": clean_text(road_name),
        "road_address": road_address,
        "normalized_road_address": normalize_address(road_address),
        "apartment_name": clean_text(apartment_name),
        "normalized_apartment_name": normalize_name(apartment_name),
        "contract_date": date_value,
        "contract_year": contract_year,
        "contract_month": contract_month,
        "area_m2": area,
        "floor": floor,
        "built_year": built_year,
        "trade_price_manwon": trade_price if transaction_type == "trade" else "",
        "deposit_manwon": deposit if transaction_type == "rent" else "",
        "monthly_rent_manwon": monthly if transaction_type == "rent" else "",
        "deal_date": date_value,
        "deal_year": contract_year,
        "deal_month": contract_month,
        "apt_name_raw": clean_text(apartment_name),
        "apt_name_normalized": normalize_name(apartment_name),
        "build_year": built_year,
        "deal_amount": trade_price if transaction_type == "trade" else "",
        "deposit_amount": deposit if transaction_type == "rent" else "",
        "monthly_rent": monthly if transaction_type == "rent" else "",
        "lat": "",
        "lng": "",
        "source_year": detect_year(Path(source_file), {"계약년월": contract_year}),
        "building_use": "아파트",
    }
    return base


def dedupe_key(row):
    return (
        row.get("transaction_type"),
        row.get("deal_type"),
        row.get("gu"),
        row.get("legal_dong"),
        row.get("road_name"),
        row.get("bonbun"),
        row.get("bubun"),
        row.get("apartment_name"),
        row.get("contract_date"),
        row.get("area_m2"),
        row.get("floor"),
        row.get("trade_price_manwon"),
        row.get("deposit_manwon"),
        row.get("monthly_rent_manwon"),
    )


def build_master():
    ensure_transaction_dirs()
    rows = []
    seen = set()
    source_stats = Counter()
    type_stats = Counter()
    skipped = Counter()

    for path, raw_rows, headers in iter_molit_sources():
        transaction_type = detect_transaction_type(path, headers)
        for raw in raw_rows:
            parsed = normalize_row(raw, path.name, transaction_type)
            if not parsed:
                skipped[path.name] += 1
                continue
            key = dedupe_key(parsed)
            if key in seen:
                continue
            seen.add(key)
            rows.append(parsed)
            source_stats[path.name] += 1
            type_stats[(parsed["transaction_type"], parsed["deal_type"])] += 1

    rows.sort(key=lambda row: (row.get("contract_date") or "", row.get("transaction_type") or ""), reverse=True)
    write_csv(TRANSACTION_MASTER_PATH, rows, MASTER_FIELDS)

    print(f"[OK] transaction_master.csv rows={len(rows)} path={TRANSACTION_MASTER_PATH}")
    if not source_stats:
        print(f"[WARNING] no MOLIT raw files found in {MOLIT_TRANSACTION_RAW_DIR}")
        print("[INFO] Put files such as trade_2024.xlsx, trade_2025.xlsx, rent_2024.xlsx in data/transactions/raw/molit/.")
        return

    print("[OK] source rows:")
    for source, count in source_stats.items():
        print(f"  - {source}: {count:,} rows")
    print("[OK] type rows:")
    for (transaction_type, deal_type), count in type_stats.items():
        print(f"  - {transaction_type}/{deal_type}: {count:,} rows")
    if skipped:
        print("[INFO] skipped rows:")
        for source, count in skipped.items():
            print(f"  - {source}: {count:,} rows")


if __name__ == "__main__":
    build_master()
